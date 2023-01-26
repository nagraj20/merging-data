from tkinter.tix import COLUMN
from PyPDF2 import PdfReader
import openpyxl
import pandas as pd

def check_cond():
    fname = 'output.xlsx'
    wb = openpyxl.load_workbook(fname)
    sheet = wb.get_sheet_by_name('Sheet1')
    for i in range(2,4):
        
        fruit_name = sheet.cell(row = i, column =2).value
        q1 =float(sheet.cell(row = i, column = 3).value)
        p1 =float(sheet.cell(row = i, column = 4).value)
        t1 =p1/q1
        q2 =float(sheet.cell(row = i, column = 5).value)
        p2 =float(sheet.cell(row = i, column = 6).value)
        t2 = p2/q2
        print('---------------------------------------------------')
        if t1 > t2:
            
            print('\n',fruit_name,'price in excel is greater per kg \n')
        elif t1 == t2:
            print( fruit_name,' price in both excel and pdf is same\n')
        else:
            print('\n',fruit_name,'price in pdf greater per kg')




wb = openpyxl.Workbook()
ws = wb.active


reader = PdfReader("book3.pdf")
page = reader.pages[0]
page_content = page.extract_text()

table_list = page_content.split('\n')

ls = []
for i in table_list:
    a=[]
    a = i.split(' ')
    ls.append(a)

tup = tuple(ls)

for i in tup:
    ws.append(i)

wb.save('fruit_convert.xlsx')

excel1 = 'fruit.xlsx'
excel2 = 'fruit_convert.xlsx'

df1 = pd.read_excel(excel1)
df2 = pd.read_excel(excel2)

merge = pd.merge(df1, df2, on = "Fruit")

merge.to_excel('output.xlsx')

check_cond()